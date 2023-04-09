from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import undetected_chromedriver as uc
import time
import sys
from openpyxl import Workbook,load_workbook
import os
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from undetected_chromedriver import Chrome, ChromeOptions


excel_file_name = "product2.xlsx"




def visibil_element(driver, by, selector, wait=20): ### web element find and search
    element = False
    if by == 'name':
        byselector = By.NAME
    if by == 'xpath':
        byselector = By.XPATH
    if by == 'css':
        byselector = By.CSS_SELECTOR
    if by == 'id':
        byselector = By.ID
    try:
        element = WebDriverWait(driver, wait).until(
            EC.visibility_of_element_located((byselector, selector)))
    except Exception as e:
        # print(e)
        element = False
    if element == False:
        pass
        # print("visibil_element not find: ", selector)
    else:
        pass
        # print(selector)
    return element

def mail_copy():
    print("Mail Copy Start")
    try:
        time.sleep(3)
        driver.get('https://temp-mail.org/en/')
        time.sleep(15)
        driver.find_element(By.XPATH, "//body").send_keys(Keys.PAGE_DOWN)
        try:
            mail_copy = visibil_element(driver,'xpath', '//button[@id="click-to-copy"]', 40)
            mail_copy.click()
            driver.execute_script("window.open('');")
        except:
            time.sleep(3)
            mail_copy = visibil_element(driver,'xpath', '//button[@id="click-to-copy"]', 40)
            mail_copy.click()
            driver.execute_script("window.open('');")
        return True
    except Exception:
        return False
        
def reg():
    print("Reg Copy Start")
    
    driver.switch_to.window(driver.window_handles[1])
    time.sleep(2)
    
    driver.get('https://supplier.coupahost.com/sessions/new?page=signup')
    business_name = visibil_element(driver,'xpath','//input[@name="business_name"]', 30)
    business_name.send_keys('Pointssoft')
                                    
    email = visibil_element(driver,'xpath','//input[@name="email"]', 30)
    email.send_keys(Keys.CONTROL + 'v')

    first = visibil_element(driver,'xpath','//input[@name="firstName"]', 30)
    first.send_keys("Alex")

    last = visibil_element(driver,'xpath','//input[@name="lastName"]', 30)
    last.send_keys("Roy")

    passwd1 = visibil_element(driver,'xpath','//input[@name="password"]', 30)
    passwd1.send_keys('gmail.com1029')

    passwd2 = visibil_element(driver,'xpath','//input[@name="confirm_password"]', 30)
    passwd2.send_keys('gmail.com1029')

    check_box = visibil_element(driver,'xpath','//input[@type="checkbox"]', 30)
    check_box.click()

    submit = visibil_element(driver,'xpath', "//button[@type='submit']", 30)
    submit.click()

    submit = visibil_element(driver,'xpath', "//button[@type='submit']", 30)
    submit.click()

def return_code():
    print("return_code Copy Start")
    
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(2)
    driver.find_element(By.XPATH, "//body").send_keys(Keys.PAGE_DOWN)
    open_mail = visibil_element(driver,'xpath', '//span[@class="bullets-ico is-active"]//parent::a[contains(@href, "https")]', 80)
    open_mail.click()
    time.sleep(2)
    driver.find_element(By.XPATH, "//body").send_keys(Keys.PAGE_DOWN)
    time.sleep(1)
    driver.find_element(By.XPATH, "//body").send_keys(Keys.PAGE_DOWN)
    code = visibil_element(driver,'xpath', '//th//h1', 30)
    code = code.text
    return code

def code_submit(mail_code):
    print("code_submit Copy Start")
    driver.switch_to.window(driver.window_handles[1])
    enter_code = visibil_element(driver,'xpath', '//input[@type="text"]', 30)
    for x in range(6):
        enter_code = driver.find_element(By.XPATH, f'(//input[@type="text"])[{x+1}]').send_keys(mail_code[x])
    submit = visibil_element(driver,'xpath', "//button[@type='submit']", 30)
    submit.click()
    time.sleep(5)
    driver.get("https://supplier.coupahost.com/connection_requests")
    search_click = visibil_element(driver,'xpath', '//button[@class="cardButton -primary searchButton s-searchButton"]',30)
    try:
        close = visibil_element(driver,'xpath', '//button[@aria-label="Close"]',5)
        close.click()
    except Exception:
        pass
    search_click.click()

def item_list(search_key):
    search_field = visibil_element(driver,'xpath', '//input[@name="customerSearchString"]' , 30)
    search_field.send_keys(search_key)
    search_field.send_keys(Keys.RETURN)
    time.sleep(5)
    items = driver.find_elements(By.XPATH,"//div[contains(@class , 'searchResults__item')]")
    item_list = []
    for item in items:
        # print(item.text)
        item_list.append(item.text)
    search_field = visibil_element(driver,'xpath', '//input[@name="customerSearchString"]' , 30)
    search_field.clear()
    return item_list

def create_workbook(excel_file_name):
    open(excel_file_name, "a")
    if os.path.getsize(excel_file_name) == 0:
        workbook = Workbook()
        spreadsheet = workbook.active
        spreadsheet.cell(row=1, column=1).value="Company name"
        workbook.save(filename=excel_file_name)

def excel_add(product_details,excel_file_name):
    workbook = load_workbook(filename=excel_file_name)
    spreadsheet = workbook.active
    ### Write
    empty_cell_row = 1
    while True:
        cell_value = spreadsheet.cell(row = empty_cell_row, column = 1).value
        if cell_value == None:
            break
        empty_cell_row += 1
    entry_column = 1
    spreadsheet.cell(row = empty_cell_row, column = entry_column).value = product_details
    workbook.save(filename=excel_file_name)
    
create_workbook(excel_file_name)

all_list = [['iu', 'iv', 'iw', 'ix', 'iy'],
 ['iz', 'ja', 'jb', 'jc', 'jd'],
 ['je', 'jf', 'jg', 'jh', 'ji'],
 ['jj', 'jk', 'jl', 'jm', 'jn'],
 ['jo', 'jp', 'jq', 'jr', 'js'],
 ['jt', 'ju', 'jv', 'jw', 'jx'],
 ['jy', 'jz', 'ka', 'kb', 'kc'],
 ['kd', 'ke', 'kf', 'kg', 'kh'],
 ['ki', 'kj', 'kk', 'kl', 'km'],
 ['kn', 'ko', 'kp', 'kq', 'kr'],
 ['ks', 'kt', 'ku', 'kv', 'kw'],
 ['kx', 'ky', 'kz', 'la', 'lb'],
 ['lc', 'ld', 'le', 'lf', 'lg'],
 ['lh', 'li', 'lj', 'lk', 'll'],
 ['lm', 'ln', 'lo', 'lp', 'lq'],
 ['lr', 'ls', 'lt', 'lu', 'lv'],
 ['lw', 'lx', 'ly', 'lz', 'ma'],
 ['mb', 'mc', 'md', 'me', 'mf'],
 ['mg', 'mh', 'mi', 'mj', 'mk'],
 ['ml', 'mm', 'mn', 'mo', 'mp'],
 ['mq', 'mr', 'ms', 'mt', 'mu'],
 ['mv', 'mw', 'mx', 'my', 'mz'],
 ['na', 'nb', 'nc', 'nd', 'ne'],
 ['nf', 'ng', 'nh', 'ni', 'nj'],
 ['nk', 'nl', 'nm', 'nn', 'no'],
 ['np', 'nq', 'nr', 'ns', 'nt'],
 ['nu', 'nv', 'nw', 'nx', 'ny'],
 ['nz', 'oa', 'ob', 'oc', 'od'],
 ['oe', 'of', 'og', 'oh', 'oi'],
 ['oj', 'ok', 'ol', 'om', 'on'],
 ['oo', 'op', 'oq', 'or', 'os'],
 ['ot', 'ou', 'ov', 'ow', 'ox'],
 ['oy', 'oz', 'pa', 'pb', 'pc'],
 ['pd', 'pe', 'pf', 'pg', 'ph'],
 ['pi', 'pj', 'pk', 'pl', 'pm'],
 ['pn', 'po', 'pp', 'pq', 'pr'],
 ['ps', 'pt', 'pu', 'pv', 'pw'],
 ['px', 'py', 'pz', 'qa', 'qb'],
 ['qc', 'qd', 'qe', 'qf', 'qg'],
 ['qh', 'qi', 'qj', 'qk', 'ql'],
 ['qm', 'qn', 'qo', 'qp', 'qq'],
 ['qr', 'qs', 'qt', 'qu', 'qv'],
 ['qw', 'qx', 'qy', 'qz', 'ra'],
 ['rb', 'rc', 'rd', 're', 'rf'],
 ['rg', 'rh', 'ri', 'rj', 'rk']]
for search_keys in all_list:
    options = ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = Chrome(options=options)
    driver.set_window_size(850,600)

    mail_copy_val = mail_copy() 
    loop_count = 1
    if mail_copy_val == False:
        for az in range(6):
            print("Trying ", az)
            mail_copy_val2 = mail_copy()
            if mail_copy_val2 == True:
                break
            loop_count += 1
    if loop_count > 5:
        print("MAy be an error found in mail copy function.")
        sys.exit()
    reg()
    mail_code = return_code()
    code_submit(mail_code)
    for search_key in search_keys:
        print("Running Item>> ", search_key)
        item = item_list(search_key)
        for i in item:
            excel_add(i,excel_file_name)
            
        print("item add success")
    driver.quit()
    print("done", search_keys)